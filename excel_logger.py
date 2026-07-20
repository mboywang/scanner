"""
Appends scanned-document records to the finance team's Excel log
(`Scanned Documents Log.xlsx`, sheet "Document Log").

If the workbook is open in Excel (Windows file lock), rows are written to a
visible fallback CSV next to it (`Scanned Documents Log (pending).csv`) and
merged into the workbook automatically as soon as it is unlocked -- so nothing
is ever lost and the pending items stay readable in the meantime.
"""
import os
import csv
import threading
from datetime import datetime

import openpyxl

from scanner_log import get_logger, activity

SHEET_NAME = "Document Log"
COLUMNS = [
    "Date Processed", "Original Filename", "Filed Filename", "Addressed To",
    "Sender Name", "Letter Date", "Due Date", "Amount Due", "Description",
    "Tasks", "Filed Location", "ClickUp Task URL", "Entity", "Subfolder",
]


class ExcelLogger:
    def __init__(self, xlsx_path):
        self.xlsx_path = xlsx_path
        self.log = get_logger()
        # Visible fallback log lives next to the workbook.
        stem = os.path.splitext(os.path.basename(xlsx_path))[0]
        self.pending_path = os.path.join(os.path.dirname(xlsx_path), f"{stem} (pending).csv")
        # Serialize all workbook writes (OCR worker + periodic flush may race).
        self._lock = threading.Lock()

    # ---- public API -----------------------------------------------------
    def append_scan(self, pdf_path, fields, filed_location="", subfolder=""):
        """Build a row from extracted fields and add it to the log (or queue it).

        filed_location/subfolder let a backfill record where a scan already lives;
        live scans leave them blank (the file stays in the main folder).
        """
        row = self._row_from_fields(pdf_path, fields or {})
        if filed_location:
            row["Filed Location"] = filed_location
        if subfolder:
            row["Subfolder"] = subfolder
        with self._lock:
            self._append_rows([row])

    def flush_pending(self):
        """Try to merge any queued rows into the workbook. Returns count flushed."""
        with self._lock:
            queued = self._read_pending()
            if not queued:
                return 0
            try:
                self._write_rows(queued)
            except PermissionError:
                return 0  # still locked; leave the fallback CSV in place
            self._clear_pending()
        activity(self.log, f"Excel: merged {len(queued)} pending row(s) from the fallback log")
        return len(queued)

    def pending_count(self):
        return len(self._read_pending())

    # ---- row building ---------------------------------------------------
    def _row_from_fields(self, pdf_path, f):
        return {
            "Date Processed": datetime.now().strftime("%m/%d/%Y"),
            "Original Filename": os.path.basename(pdf_path),
            "Filed Filename": "",             # finance fills after filing
            "Addressed To": f.get("addressed_to", ""),
            "Sender Name": f.get("sender_name", ""),
            "Letter Date": f.get("letter_date", ""),
            "Due Date": f.get("due_date", ""),
            "Amount Due": f.get("amount_due", ""),
            "Description": f.get("description", ""),
            "Tasks": f.get("tasks", ""),
            "Filed Location": "",             # finance fills
            "ClickUp Task URL": "",           # finance fills
            "Entity": f.get("entity", ""),
            "Subfolder": f.get("subfolder", ""),
        }

    # ---- writing (callers must hold self._lock) -------------------------
    def _append_rows(self, rows):
        # Drain the fallback queue first so ordering is roughly preserved.
        all_rows = self._read_pending() + rows
        try:
            self._write_rows(all_rows)
            self._clear_pending()
            for r in rows:
                activity(self.log,
                         f"Excel: logged '{r['Original Filename']}' "
                         f"(sender='{r['Sender Name']}', amount='{r['Amount Due']}')")
        except PermissionError:
            # Workbook is open in Excel - write to the visible fallback CSV.
            self._queue(rows)
            activity(self.log,
                     f"Excel locked (open in Excel?) - wrote {len(rows)} row(s) to "
                     f"'{os.path.basename(self.pending_path)}'; will merge when unlocked")

    def _write_rows(self, rows):
        """Open the workbook, append rows to the Document Log sheet, save."""
        if os.path.exists(self.xlsx_path):
            wb = openpyxl.load_workbook(self.xlsx_path)
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append(COLUMNS)

        header = [c.value for c in ws[1]]
        col_index = {name: i for i, name in enumerate(header) if name in COLUMNS}
        for row in rows:
            values = [""] * len(header)
            for name, idx in col_index.items():
                values[idx] = row.get(name, "")
            ws.append(values)
        wb.save(self.xlsx_path)

    # ---- fallback CSV queue ---------------------------------------------
    def _queue(self, rows):
        exists = os.path.exists(self.pending_path)
        try:
            with open(self.pending_path, "a", newline="", encoding="utf-8-sig") as fh:
                writer = csv.DictWriter(fh, fieldnames=COLUMNS)
                if not exists:
                    writer.writeheader()
                for r in rows:
                    writer.writerow({c: r.get(c, "") for c in COLUMNS})
        except Exception as e:
            # Last resort: never crash a scan over logging.
            self.log.exception("Failed to write fallback pending CSV: %s", e)

    def _read_pending(self):
        if not os.path.exists(self.pending_path):
            return []
        try:
            with open(self.pending_path, "r", newline="", encoding="utf-8-sig") as fh:
                return [dict(r) for r in csv.DictReader(fh)]
        except Exception:
            return []

    def _clear_pending(self):
        try:
            if os.path.exists(self.pending_path):
                os.remove(self.pending_path)
        except Exception:
            pass
